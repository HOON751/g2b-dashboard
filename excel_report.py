"""
엑셀 양식 보고서 생성 모듈 (excel_report.py) - v3

특징:
- 3개 시트로 분리:
  · 시트 1: 발주리스트 (전체 발주 내역)
  · 시트 2: 권역별 업체 발주량 (집계)
  · 시트 3: 경쟁사별 M/S
- 검색 기간의 월만 표시
- 각 월에 실제 데이터 건수만큼만 행 할당
- 우수제품 표시는 우수제품 계약이 1건이라도 있는 업체에만 정확히
"""

import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


# ──────────────── 스타일 ────────────────
THIN = Side(border_style="thin", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
SECTION_FILL = PatternFill("solid", fgColor="FFF2CC")
SUBTOTAL_FILL = PatternFill("solid", fgColor="EAEAEA")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")

EXCL_FONT = Font(name="맑은 고딕", size=10, color="C00000", bold=True)
HEADER_FONT = Font(name="맑은 고딕", size=10, bold=True)
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name="맑은 고딕", size=10, color="555555")
GEN_FONT = Font(name="맑은 고딕", size=9, color="888888")
NORMAL_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def _short_region(v):
    if not v:
        return ""
    s = str(v).split()[0] if v else ""
    region_map = {
        "서울특별시": "서울", "부산광역시": "부산", "대구광역시": "대구",
        "인천광역시": "인천", "광주광역시": "광주", "대전광역시": "대전",
        "울산광역시": "울산", "세종특별자치시": "세종",
        "경기도": "경기", "강원도": "강원", "강원특별자치도": "강원",
        "충청북도": "충북", "충청남도": "충남",
        "전라북도": "전북", "전북특별자치도": "전북", "전라남도": "전남",
        "경상북도": "경북", "경상남도": "경남",
        "제주특별자치도": "제주", "제주도": "제주",
    }
    return region_map.get(s, s[:2] if s else "")


def _is_education(govdiv):
    return "교육" in str(govdiv or "")


def _months_in_range(bgn_str, end_str):
    try:
        b = pd.to_datetime(bgn_str)
        e = pd.to_datetime(end_str)
    except Exception:
        return []
    months = []
    cur = pd.Timestamp(year=b.year, month=b.month, day=1)
    last = pd.Timestamp(year=e.year, month=e.month, day=1)
    while cur <= last:
        months.append((cur.year, cur.month))
        if cur.month == 12:
            cur = pd.Timestamp(year=cur.year + 1, month=1, day=1)
        else:
            cur = pd.Timestamp(year=cur.year, month=cur.month + 1, day=1)
    return months


def _set(ws, r, c, value=None, font=None, fill=None, align=None,
         border=BORDER_ALL, number_format=None):
    cell = ws.cell(r, c)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
                cell = ws.cell(mr.min_row, mr.min_col)
                break
    if value is not None: cell.value = value
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    if number_format: cell.number_format = number_format
    return cell


def _add_title_block(ws, title, term, bgn, end):
    """각 시트 상단에 일관된 제목 블록 추가. 반환: 다음 사용 가능 행"""
    ws.row_dimensions[1].height = 32
    _set(ws, 1, 1, value=title, font=TITLE_FONT, fill=TITLE_FILL,
         align=Alignment(horizontal="center", vertical="center"),
         border=None)
    _set(ws, 2, 1, value=f"품목: {term}    조회기간: {bgn} ~ {end}",
         font=SUBTITLE_FONT,
         align=Alignment(horizontal="left", vertical="center"),
         border=None)
    _set(ws, 3, 1,
         value=f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}    (단위: 원, 부가세 별도)",
         font=GEN_FONT,
         align=Alignment(horizontal="left", vertical="center"),
         border=None)
    return 5  # 빈 줄 하나 두고 5행부터 시작


# ════════════════ 시트 1: 발주리스트 ════════════════
def _build_sheet1(wb, data_rows, months_to_show, term, bgn, end):
    ws = wb.create_sheet("발주리스트")

    # 컬럼 너비
    widths = {"A": 7, "B": 12, "C": 22, "D": 12, "E": 26, "F": 8, "G": 40, "H": 16, "I": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 제목 블록: A~I 9개 컬럼이라 1~9 머지
    ws.merge_cells("A1:I1")
    ws.merge_cells("A2:I2")
    ws.merge_cells("A3:I3")
    start_row = _add_title_block(ws, "발주리스트 (전체 내역)", term, bgn, end)

    # 헤더
    headers = ["월별", "계약일", "업체명", "구분", "수요기관", "지역",
               "계약건명", "금액", "비고"]
    for c, h in enumerate(headers, start=1):
        _set(ws, start_row, c, value=h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    ws.row_dimensions[start_row].height = 28

    current_row = start_row + 1
    data_start = current_row

    if not months_to_show:
        _set(ws, current_row, 1, value="(해당 조건의 데이터 없음)", align=CENTER)
        for c in range(1, 10):
            _set(ws, current_row, c, border=BORDER_ALL)
        ws.merge_cells(f"A{current_row}:I{current_row}")
        current_row += 1
    else:
        for (yr, mo) in months_to_show:
            month_rows = [r for r in data_rows if r["year"] == yr and r["month"] == mo]
            if not month_rows:
                continue
            month_start = current_row
            for r in month_rows:
                _set(ws, current_row, 1, value=f"{mo}월", font=BOLD_FONT,
                     fill=SECTION_FILL, align=CENTER)
                _set(ws, current_row, 2, value=r["date"].date(),
                     font=NORMAL_FONT, align=CENTER, number_format="yyyy-mm-dd")
                _set(ws, current_row, 3, value=r["업체명"], font=NORMAL_FONT, align=LEFT)
                _set(ws, current_row, 4, value=r["구분"], font=NORMAL_FONT, align=CENTER)
                _set(ws, current_row, 5, value=r["수요기관"], font=NORMAL_FONT, align=LEFT)
                _set(ws, current_row, 6, value=r["지역"], font=NORMAL_FONT, align=CENTER)
                _set(ws, current_row, 7, value=r["계약건명"], font=NORMAL_FONT, align=LEFT)
                _set(ws, current_row, 8, value=int(r["금액"]),
                     font=NORMAL_FONT, align=RIGHT, number_format="#,##0")
                _set(ws, current_row, 9, value=r["비고"], font=NORMAL_FONT, align=CENTER)
                current_row += 1
            month_end = current_row - 1
            if month_end > month_start:
                try:
                    ws.merge_cells(f"A{month_start}:A{month_end}")
                except Exception:
                    pass

    # 합계
    data_end = current_row - 1
    _set(ws, current_row, 1, value="합계", font=BOLD_FONT,
         fill=SUBTOTAL_FILL, align=CENTER)
    for c in range(2, 8):
        _set(ws, current_row, c, fill=SUBTOTAL_FILL, border=BORDER_ALL)
    ws.merge_cells(f"A{current_row}:G{current_row}")
    if data_end >= data_start:
        _set(ws, current_row, 8,
             value=f"=SUM(H{data_start}:H{data_end})",
             font=BOLD_FONT, fill=SUBTOTAL_FILL,
             align=RIGHT, number_format="#,##0")
    else:
        _set(ws, current_row, 8, value=0,
             font=BOLD_FONT, fill=SUBTOTAL_FILL,
             align=RIGHT, number_format="#,##0")
    _set(ws, current_row, 9, fill=SUBTOTAL_FILL, border=BORDER_ALL)

    # 데이터 영역 행 높이
    for r in range(start_row + 1, current_row):
        ws.row_dimensions[r].height = 22

    # 자동 필터
    if data_end >= data_start:
        ws.auto_filter.ref = f"A{start_row}:I{data_end}"
    # 윗부분 고정 (제목+헤더)
    ws.freeze_panes = f"A{start_row + 1}"


# ════════════════ 시트 2: 권역별 업체 발주량 ════════════════
def _build_sheet2(wb, top_corps, term, bgn, end):
    ws = wb.create_sheet("권역별 업체 발주량")

    widths = {"A": 6, "B": 26, "C": 14, "D": 16, "E": 16, "F": 16, "G": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:G1")
    ws.merge_cells("A2:G2")
    ws.merge_cells("A3:G3")
    start_row = _add_title_block(ws, "1) 권역별 업체 발주량", term, bgn, end)

    headers = ["NO.", "업체명", "소재지", "합계", "지자체외", "교육기관", "비고"]
    for c, h in enumerate(headers, start=1):
        _set(ws, start_row, c, value=h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    ws.row_dimensions[start_row].height = 28

    current_row = start_row + 1
    data_start = current_row

    if not top_corps:
        _set(ws, current_row, 1, value="(데이터 없음)", align=CENTER)
        for c in range(1, 8):
            _set(ws, current_row, c, border=BORDER_ALL)
        ws.merge_cells(f"A{current_row}:G{current_row}")
        current_row += 1
    else:
        for i, (corp_name, stats) in enumerate(top_corps[:30], start=1):
            _set(ws, current_row, 1, value=i, font=NORMAL_FONT, align=CENTER)
            _set(ws, current_row, 2, value=corp_name, font=NORMAL_FONT, align=LEFT)
            _set(ws, current_row, 3, value="", font=NORMAL_FONT, align=CENTER)
            _set(ws, current_row, 4, value=int(stats["total"]),
                 font=NORMAL_FONT, align=RIGHT, number_format="#,##0")
            _set(ws, current_row, 5, value=int(stats["non_edu"]),
                 font=NORMAL_FONT, align=RIGHT, number_format="#,##0")
            _set(ws, current_row, 6, value=int(stats["edu"]),
                 font=NORMAL_FONT, align=RIGHT, number_format="#,##0")
            if stats["has_excl"]:
                _set(ws, current_row, 7, value="조달우수",
                     font=EXCL_FONT, align=CENTER)
            else:
                _set(ws, current_row, 7, value="", font=NORMAL_FONT, align=CENTER)
            current_row += 1

    # 합계
    data_end = current_row - 1
    _set(ws, current_row, 1, value="합계", font=BOLD_FONT,
         fill=SUBTOTAL_FILL, align=CENTER)
    _set(ws, current_row, 2, fill=SUBTOTAL_FILL, border=BORDER_ALL)
    _set(ws, current_row, 3, fill=SUBTOTAL_FILL, border=BORDER_ALL)
    ws.merge_cells(f"A{current_row}:C{current_row}")
    if data_end >= data_start:
        for col, letter in [(4, "D"), (5, "E"), (6, "F")]:
            _set(ws, current_row, col,
                 value=f"=SUM({letter}{data_start}:{letter}{data_end})",
                 font=BOLD_FONT, fill=SUBTOTAL_FILL,
                 align=RIGHT, number_format="#,##0")
    else:
        for col in [4, 5, 6]:
            _set(ws, current_row, col, value=0,
                 font=BOLD_FONT, fill=SUBTOTAL_FILL,
                 align=RIGHT, number_format="#,##0")
    _set(ws, current_row, 7, fill=SUBTOTAL_FILL, border=BORDER_ALL)

    for r in range(start_row + 1, current_row):
        ws.row_dimensions[r].height = 22

    if data_end >= data_start:
        ws.auto_filter.ref = f"A{start_row}:G{data_end}"
    ws.freeze_panes = f"A{start_row + 1}"


# ════════════════ 시트 3: 경쟁사별 M/S ════════════════
def _build_sheet3(wb, top_corps, total_all, total_nonedu, total_edu, term, bgn, end):
    ws = wb.create_sheet("경쟁사별 M_S")

    widths = {"A": 6, "B": 26, "C": 14,
              "D": 16, "E": 10,
              "F": 16, "G": 10,
              "H": 16, "I": 10,
              "J": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:J1")
    ws.merge_cells("A2:J2")
    ws.merge_cells("A3:J3")
    start_row = _add_title_block(ws, "2) 경쟁사별 M/S", term, bgn, end)

    # 헤더 - 2단 (합계/M/S, 지자체외/M/S, 교육기관/M/S)
    # 1행 헤더
    _set(ws, start_row, 1, value="NO.", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, start_row, 2, value="업체명", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, start_row, 3, value="소재지", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, start_row, 4, value="합계", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, start_row, 5, value="M/S", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, start_row, 6, value="지자체외", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, start_row, 7, value="M/S", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, start_row, 8, value="교육기관", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, start_row, 9, value="M/S", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, start_row, 10, value="비고", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    ws.row_dimensions[start_row].height = 28

    current_row = start_row + 1
    data_start = current_row

    if not top_corps:
        _set(ws, current_row, 1, value="(데이터 없음)", align=CENTER)
        for c in range(1, 11):
            _set(ws, current_row, c, border=BORDER_ALL)
        ws.merge_cells(f"A{current_row}:J{current_row}")
        current_row += 1
    else:
        for i, (corp_name, stats) in enumerate(top_corps[:30], start=1):
            _set(ws, current_row, 1, value=i, font=NORMAL_FONT, align=CENTER)
            _set(ws, current_row, 2, value=corp_name, font=NORMAL_FONT, align=LEFT)
            _set(ws, current_row, 3, value="", font=NORMAL_FONT, align=CENTER)

            _set(ws, current_row, 4, value=int(stats["total"]),
                 font=NORMAL_FONT, align=RIGHT, number_format="#,##0")
            ms_total = stats["total"] / total_all if total_all else 0
            _set(ws, current_row, 5, value=ms_total,
                 font=NORMAL_FONT, align=CENTER, number_format="0.0%")

            _set(ws, current_row, 6, value=int(stats["non_edu"]),
                 font=NORMAL_FONT, align=RIGHT, number_format="#,##0")
            ms_nonedu = stats["non_edu"] / total_nonedu if total_nonedu else 0
            _set(ws, current_row, 7, value=ms_nonedu,
                 font=NORMAL_FONT, align=CENTER, number_format="0.0%")

            _set(ws, current_row, 8, value=int(stats["edu"]),
                 font=NORMAL_FONT, align=RIGHT, number_format="#,##0")
            ms_edu = stats["edu"] / total_edu if total_edu else 0
            _set(ws, current_row, 9, value=ms_edu,
                 font=NORMAL_FONT, align=CENTER, number_format="0.0%")

            if stats["has_excl"]:
                _set(ws, current_row, 10, value="조달우수",
                     font=EXCL_FONT, align=CENTER)
            else:
                _set(ws, current_row, 10, value="", font=NORMAL_FONT, align=CENTER)
            current_row += 1

    # 합계
    data_end = current_row - 1
    _set(ws, current_row, 1, value="합계", font=BOLD_FONT,
         fill=SUBTOTAL_FILL, align=CENTER)
    _set(ws, current_row, 2, fill=SUBTOTAL_FILL, border=BORDER_ALL)
    _set(ws, current_row, 3, fill=SUBTOTAL_FILL, border=BORDER_ALL)
    ws.merge_cells(f"A{current_row}:C{current_row}")
    if data_end >= data_start:
        _set(ws, current_row, 4,
             value=f"=SUM(D{data_start}:D{data_end})",
             font=BOLD_FONT, fill=SUBTOTAL_FILL, align=RIGHT, number_format="#,##0")
        _set(ws, current_row, 5, value=1.0,
             font=BOLD_FONT, fill=SUBTOTAL_FILL, align=CENTER, number_format="0.0%")
        _set(ws, current_row, 6,
             value=f"=SUM(F{data_start}:F{data_end})",
             font=BOLD_FONT, fill=SUBTOTAL_FILL, align=RIGHT, number_format="#,##0")
        _set(ws, current_row, 7, value=1.0,
             font=BOLD_FONT, fill=SUBTOTAL_FILL, align=CENTER, number_format="0.0%")
        _set(ws, current_row, 8,
             value=f"=SUM(H{data_start}:H{data_end})",
             font=BOLD_FONT, fill=SUBTOTAL_FILL, align=RIGHT, number_format="#,##0")
        _set(ws, current_row, 9, value=1.0,
             font=BOLD_FONT, fill=SUBTOTAL_FILL, align=CENTER, number_format="0.0%")
    else:
        for col in [4, 5, 6, 7, 8, 9]:
            _set(ws, current_row, col, value=0,
                 font=BOLD_FONT, fill=SUBTOTAL_FILL,
                 align=RIGHT if col % 2 == 0 else CENTER, number_format="#,##0")
    _set(ws, current_row, 10, fill=SUBTOTAL_FILL, border=BORDER_ALL)

    for r in range(start_row + 1, current_row):
        ws.row_dimensions[r].height = 22

    if data_end >= data_start:
        ws.auto_filter.ref = f"A{start_row}:J{data_end}"
    ws.freeze_panes = f"A{start_row + 1}"


# ════════════════ 메인 함수 ════════════════
def build_excel_report(df, term, bgn, end):
    """검색된 DataFrame을 3개 시트 엑셀로 만들어 BytesIO 반환"""
    wb = Workbook()
    # 기본 시트 제거
    default = wb.active
    wb.remove(default)

    # 데이터 정리
    data_rows = []
    if "계약일자_dt" in df.columns and len(df) > 0:
        sdf = df.sort_values("계약일자_dt").copy()
        for _, it in sdf.iterrows():
            dt = it.get("계약일자_dt")
            if pd.isna(dt):
                continue
            data_rows.append({
                "month": dt.month,
                "year": dt.year,
                "date": dt,
                "업체명": it.get("업체명") or "",
                "구분": it.get("수요기관구분") or it.get("소관구분") or "",
                "수요기관": it.get("수요기관") or "",
                "지역": _short_region(it.get("수요기관지역")),
                "계약건명": it.get("계약명") or "",
                "금액": float(it.get("공급금액") or 0),
                "비고": it.get("계약방법") or "",
                "우수제품": str(it.get("우수제품여부") or "").upper() == "Y",
            })

    # 표시할 월 결정
    months_in_range = _months_in_range(bgn, end)
    months_with_data = sorted(set((r["year"], r["month"]) for r in data_rows))
    months_to_show = [m for m in months_in_range if m in months_with_data] or months_with_data

    # 업체별 집계
    corp_stats = {}
    total_all = 0
    total_nonedu = 0
    total_edu = 0
    for r in data_rows:
        c = r["업체명"]
        if not c:
            continue
        s = corp_stats.setdefault(c, {"total": 0, "edu": 0, "non_edu": 0, "has_excl": False})
        s["total"] += r["금액"]
        total_all += r["금액"]
        if _is_education(r["구분"]):
            s["edu"] += r["금액"]
            total_edu += r["금액"]
        else:
            s["non_edu"] += r["금액"]
            total_nonedu += r["금액"]
        if r["우수제품"]:
            s["has_excl"] = True
    top_corps = sorted(corp_stats.items(), key=lambda x: -x[1]["total"])

    # 3개 시트 생성
    _build_sheet1(wb, data_rows, months_to_show, term, bgn, end)
    _build_sheet2(wb, top_corps, term, bgn, end)
    _build_sheet3(wb, top_corps, total_all, total_nonedu, total_edu, term, bgn, end)

    # 첫 시트가 활성으로 보이도록
    wb.active = 0

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf